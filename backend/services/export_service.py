import io
import calendar
from collections import defaultdict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _cat_map(categories: list) -> dict:
    """Return {id: name} from a list of Category ORM objects or dicts."""
    result = {}
    for c in categories:
        if hasattr(c, "id"):
            result[c.id] = c.name
        else:
            result[c["id"]] = c["name"]
    return result


def _budget_limit_map(budgets: list) -> dict:
    """Return {category_id: monthly_limit}."""
    result = {}
    for b in budgets:
        if hasattr(b, "category_id"):
            result[b.category_id] = b.monthly_limit
        else:
            result[b["category_id"]] = b["monthly_limit"]
    return result


def _expense_field(expense, field: str):
    """Get a field from an ORM object or dict."""
    return getattr(expense, field, None) if hasattr(expense, field) else expense.get(field)


# ─────────────────────────────────────────────────────────────────────────────
# 1. export_to_csv
# ─────────────────────────────────────────────────────────────────────────────

def export_to_csv(expenses: list, categories: list) -> io.StringIO:
    cat_names = _cat_map(categories)

    rows = []
    for exp in expenses:
        cat_id = _expense_field(exp, "category_id")
        rows.append({
            "Date": str(_expense_field(exp, "date") or ""),
            "Merchant": _expense_field(exp, "merchant") or "",
            "Description": _expense_field(exp, "description") or "",
            "Category": cat_names.get(cat_id, "Uncategorized"),
            "Amount": _expense_field(exp, "amount"),
            "Notes": _expense_field(exp, "notes") or "",
            "Is Recurring": "Yes" if _expense_field(exp, "is_recurring") else "No",
        })

    df = pd.DataFrame(rows, columns=["Date", "Merchant", "Description", "Category", "Amount", "Notes", "Is Recurring"])
    buf = io.StringIO()
    df.to_csv(buf, index=False)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────────────────
# 2. export_to_excel
# ─────────────────────────────────────────────────────────────────────────────

def export_to_excel(
    expenses: list,
    budgets: list,
    categories: list,
    month: int,
    year: int,
) -> io.BytesIO:
    wb = Workbook()
    cat_names = _cat_map(categories)
    budget_limits = _budget_limit_map(budgets)

    # ── Shared styles ────────────────────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    alt_fill = PatternFill("solid", fgColor="F2F9FF")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    def style_header_row(ws, row_num: int, num_cols: int):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

    def auto_width(ws):
        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 50)

    # ── SHEET 1: Transactions ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Transactions"

    headers1 = ["Date", "Merchant", "Description", "Category", "Amount", "Notes"]
    ws1.append(headers1)
    style_header_row(ws1, 1, len(headers1))

    total_spend = 0.0
    top_cat_totals: dict[str, float] = defaultdict(float)

    for i, exp in enumerate(expenses, start=2):
        cat_id = _expense_field(exp, "category_id")
        cat_name = cat_names.get(cat_id, "Uncategorized")
        amount = float(_expense_field(exp, "amount") or 0)
        total_spend += amount
        top_cat_totals[cat_name] += amount

        row = [
            str(_expense_field(exp, "date") or ""),
            _expense_field(exp, "merchant") or "",
            _expense_field(exp, "description") or "",
            cat_name,
            amount,
            _expense_field(exp, "notes") or "",
        ]
        ws1.append(row)

        fill = alt_fill if i % 2 == 0 else white_fill
        for col in range(1, len(headers1) + 1):
            cell = ws1.cell(row=i, column=col)
            cell.fill = fill
            cell.border = thin_border
            if col == 5:  # Amount
                cell.number_format = '"$"#,##0.00'

    auto_width(ws1)

    # ── SHEET 2: Summary ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")

    headers2 = ["Category", "Total Spent", "% of Total", "Budget Limit", "Status"]
    ws2.append(headers2)
    style_header_row(ws2, 1, len(headers2))

    for i, (cat_name, cat_total) in enumerate(
        sorted(top_cat_totals.items(), key=lambda x: x[1], reverse=True), start=2
    ):
        pct = (cat_total / total_spend * 100) if total_spend else 0
        # Find category_id by name reverse lookup
        cat_id = next((cid for cid, name in cat_names.items() if name == cat_name), None)
        limit = budget_limits.get(cat_id) if cat_id else None
        status = ""
        if limit:
            status = "Over" if cat_total > limit else "Under"

        row = [cat_name, cat_total, f"{pct:.1f}%", limit or "", status]
        ws2.append(row)

        fill = alt_fill if i % 2 == 0 else white_fill
        for col in range(1, len(headers2) + 1):
            cell = ws2.cell(row=i, column=col)
            cell.fill = fill
            cell.border = thin_border
            if col == 2:
                cell.number_format = '"$"#,##0.00'
            if col == 4 and limit:
                cell.number_format = '"$"#,##0.00'
            if col == 5 and status == "Over":
                cell.font = Font(color="C00000", bold=True)
            elif col == 5 and status == "Under":
                cell.font = Font(color="375623", bold=True)

    auto_width(ws2)

    # ── SHEET 3: Overview ────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Overview")

    month_name = calendar.month_name[month]
    title_text = f"ExpenseIQ Monthly Report — {month_name} {year}"

    ws3.merge_cells("A1:B1")
    title_cell = ws3["A1"]
    title_cell.value = title_text
    title_cell.font = Font(bold=True, size=14, color="1F4E79")
    title_cell.alignment = center_align

    days_in_month = calendar.monthrange(year, month)[1]
    avg_per_day = total_spend / days_in_month if days_in_month else 0
    top_category = max(top_cat_totals, key=top_cat_totals.get) if top_cat_totals else "N/A"
    top_category_amount = top_cat_totals.get(top_category, 0)

    overview_rows = [
        ("", ""),  # spacer
        ("Total Spend", total_spend),
        ("Number of Transactions", len(expenses)),
        ("Average Per Day", avg_per_day),
        ("Top Category", f"{top_category} (${top_category_amount:,.0f})"),
    ]

    for row_data in overview_rows:
        ws3.append(list(row_data))

    # Style label column bold
    for row in ws3.iter_rows(min_row=3, max_row=ws3.max_row, min_col=1, max_col=2):
        label_cell, value_cell = row
        label_cell.font = Font(bold=True)
        label_cell.border = thin_border
        value_cell.border = thin_border
        if isinstance(value_cell.value, float):
            value_cell.number_format = '"$"#,##0.00'

    auto_width(ws3)

    # ── Serialize ────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
